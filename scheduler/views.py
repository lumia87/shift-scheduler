# scheduler/views.py
from django.shortcuts import render
from .scheduler_logic3 import generate_shift_schedule
from django.http import HttpResponse
import json
from openpyxl import Workbook # type: ignore
from openpyxl.styles import PatternFill # type: ignore

def parse_fixed_shifts(fixed_shifts_str_list):
    """
    Chuyển đổi danh sách chuỗi ca cố định thành dạng dictionary.
    Ví dụ: ["1,g1,s; 1,g2,c; 1,g3,d"] -> {0: {0: 's', 1: 'c', 2: 'd'}}
    """
    fixed_shifts = {}
    fixed_shifts_list=fixed_shifts_str_list[0].split(";")
    for entry in fixed_shifts_list:
        entry = entry.strip()
        if entry:
            day, gr_id, shift = entry.split(",")
            day = int(day)
            gr_id =int(gr_id[1:])
            if shift:  # Chỉ thêm nếu shift không rỗng
                if day not in fixed_shifts:
                    fixed_shifts[day] = {}
                fixed_shifts[day][gr_id] = shift
    return fixed_shifts

def json_to_excel(request):
    # Lấy JSON từ session
    if 'schedule_json' in request.session:
        json_data_str, note_fixed_shifts = request.session.get('schedule_json')
        json_data = json.loads(json_data_str)
        # Định nghĩa màu sắc cho các ca trực
        shift_colors = {
            's': 'd4edda',  # Green for morning shift
            'c': 'fff3cd',  # Blue for afternoon shift
            'd': 'f8d7da',  # Red for night shift
            'x': 'FFFFFF'   # White for day off
        }


        wb = Workbook()
        ws = wb.active
        ws.title = "Shift Schedule"
        # Thêm thông tin tháng và năm
        year = json_data.get('year')
        month = json_data.get('month')
        ws.append([f'Lịch trực tháng {month}/{year}'])
        ws.append(["Ca trực cố định:" + note_fixed_shifts])
        ws.append([])  # Dòng trống

        # Tạo tiêu đề cột
        headers = ['STT', 'Tên'] + [day['day'] for day in json_data['teams'][0]['days']] + ['T.S', 'T.C', 'T.Đ', 'Tổng']
        ws.append(headers)
        for team_idx, team in enumerate(json_data['teams']):
            for member_idx, member in enumerate(team['members']):
                row = [f"{team_idx + 1}.{member_idx + 1}", member]
                for day in team['days']:
                    row.append(day['shift'].upper())
                row += [team['shifts']['s'], team['shifts']['c'], team['shifts']['d'], team['total_shift']]
                ws.append(row)

                # Tô màu cho các ca trực
                for col_idx, day in enumerate(team['days'], start=3):
                    cell = ws.cell(row=len(ws['A']), column=col_idx)
                    cell.fill = PatternFill(start_color=shift_colors[day['shift']], end_color=shift_colors[day['shift']], fill_type="solid")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=shift_schedule.xlsx'
        wb.save(response)
        return response
    return HttpResponse("Không có dữ liệu trên session")


def index(request):
    if request.method == 'POST':
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        max_hours_per_week = int(request.POST["max_hours_per_week"])  if request.POST["max_hours_per_week"] else 40

        teams = [request.POST.get(f'team{i}').split(',') for i in range(1, 7) if request.POST.get(f'team{i}')] #chỉ lấy các teams có giá trị
        
        fixed_shifts_input = request.POST.get('fixed_shift')
        fixed_shifts = {}
        for item in fixed_shifts_input.split(';'):
            day, group, shift = item.split(',')
            day = int(day)
            group = int(group[1:])  # Bỏ "g" ở đầu và chuyển đổi sang số
            if day not in fixed_shifts:
                fixed_shifts[day] = {}
            fixed_shifts[day][group] = shift

        # Gọi hàm generate_shift_schedule và trả về JSON
        result_json = generate_shift_schedule(year, month, teams, fixed_shifts, max_hours_per_week)
        
          # Lưu JSON vào session
        request.session['schedule_json'] = result_json, fixed_shifts_input

        # Parse kết quả JSON sang dictionary
        result = json.loads(result_json)

        # Trả về template với dữ liệu JSON
        return render(request, 'scheduler/result.html', {'json_data': result, 'month':month, 'year':year})    

    return render(request, 'scheduler/index.html')